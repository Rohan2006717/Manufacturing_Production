"""
Manufacturing Production Line Optimization

Simulated analysis of a 5-stage manufacturing production line.

Production Flow:
Raw Material -> Cutting -> Machining -> Assembly -> Inspection -> Packing
"""

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

# PROJECT CONFIGURATION

RANDOM_SEED = 42

np.random.seed(RANDOM_SEED)

PLANNED_PRODUCTION_TIME = 480  # minutes per day
NUMBER_OF_DAYS = 30



# PRODUCTION LINE


WORKSTATIONS = {

    "Cutting": {
        "ideal_cycle_time": 0.90,
        "downtime_range": (20, 40),
        "defect_rate_range": (0.010, 0.025)
    },

    "Machining": {
        "ideal_cycle_time": 2.40,
        "downtime_range": (35, 65),
        "defect_rate_range": (0.015, 0.030)
    },

    "Assembly": {
        "ideal_cycle_time": 1.60,
        "downtime_range": (20, 40),
        "defect_rate_range": (0.010, 0.025)
    },

    "Inspection": {
        "ideal_cycle_time": 1.00,
        "downtime_range": (15, 30),
        "defect_rate_range": (0.005, 0.015)
    },

    "Packing": {
        "ideal_cycle_time": 0.80,
        "downtime_range": (10, 25),
        "defect_rate_range": (0.005, 0.015)
    }
}



# RESULTS DIRECTORY


RESULTS_DIR = "results"

os.makedirs(RESULTS_DIR, exist_ok=True)


# CAPACITY CALCULATION

def calculate_theoretical_capacity(cycle_time):
    """
    Calculate the theoretical daily production capacity.

    Capacity = Planned Production Time / Ideal Cycle Time
    """

    return PLANNED_PRODUCTION_TIME / cycle_time

# ============================================================
# PRODUCTION DATA GENERATION
# ============================================================

def generate_production_data(
    cycle_time_reduction=0.0,
    downtime_reduction=0.0,
    target_workstation=None
):
    """
    Generate simulated production data.

    Optional optimization parameters allow us to modify
    the cycle time and downtime of a selected workstation.

    Parameters
    ----------
    cycle_time_reduction : float
        Percentage reduction in cycle time.
        Example: 0.10 = 10% reduction.

    downtime_reduction : float
        Percentage reduction in downtime.
        Example: 0.20 = 20% reduction.

    target_workstation : str
        Workstation to which the improvements are applied.
    """

    production_records = []

    for day in range(1, NUMBER_OF_DAYS + 1):

        for workstation, parameters in WORKSTATIONS.items():

            # ------------------------------------------------
            # 1. Determine cycle time
            # ------------------------------------------------

            ideal_cycle_time = parameters["ideal_cycle_time"]

            if workstation == target_workstation:

                ideal_cycle_time *= (
                    1 - cycle_time_reduction
                )

            # ------------------------------------------------
            # 2. Generate downtime
            # ------------------------------------------------

            min_downtime, max_downtime = (
                parameters["downtime_range"]
            )

            downtime = np.random.uniform(
                min_downtime,
                max_downtime
            )

            # Apply downtime improvement
            if workstation == target_workstation:

                downtime *= (
                    1 - downtime_reduction
                )

            downtime = round(downtime, 2)

            # ------------------------------------------------
            # 3. Calculate operating time
            # ------------------------------------------------

            operating_time = (
                PLANNED_PRODUCTION_TIME - downtime
            )

            # ------------------------------------------------
            # 4. Calculate theoretical production
            # ------------------------------------------------

            theoretical_capacity = (
                operating_time / ideal_cycle_time
            )

            # ------------------------------------------------
            # 5. Production variation
            # ------------------------------------------------

            performance_factor = np.random.uniform(
                0.92,
                0.99
            )

            total_count = (
                theoretical_capacity
                * performance_factor
            )

            total_count = int(
                round(total_count)
            )

            # ------------------------------------------------
            # 6. Generate defect rate
            # ------------------------------------------------

            min_defect_rate, max_defect_rate = (
                parameters["defect_rate_range"]
            )

            defect_rate = np.random.uniform(
                min_defect_rate,
                max_defect_rate
            )

            rejected_count = int(
                round(total_count * defect_rate)
            )

            # ------------------------------------------------
            # 7. Calculate good units
            # ------------------------------------------------

            good_count = (
                total_count - rejected_count
            )

            # ------------------------------------------------
            # 8. Store record
            # ------------------------------------------------

            production_records.append({
                "Day": day,
                "Workstation": workstation,
                "Planned Production Time":
                    PLANNED_PRODUCTION_TIME,
                "Downtime":
                    downtime,
                "Operating Time":
                    round(operating_time, 2),
                "Ideal Cycle Time":
                    ideal_cycle_time,
                "Total Count":
                    total_count,
                "Good Count":
                    good_count,
                "Rejected Count":
                    rejected_count
            })

    production_df = pd.DataFrame(
        production_records
    )

    return production_df

# ============================================================
# KPI CALCULATIONS
# ============================================================

def calculate_kpis(production_df):
    """
    Calculate manufacturing KPIs for every production record.

    KPIs:
    - Availability
    - Performance
    - Quality
    - OEE
    - Throughput
    - Capacity Utilization
    """

    df = production_df.copy()

    # --------------------------------------------------------
    # Availability
    # --------------------------------------------------------

    df["Availability"] = (
        df["Operating Time"]
        / df["Planned Production Time"]
    )

    # --------------------------------------------------------
    # Performance
    # --------------------------------------------------------

    df["Performance"] = (
        df["Ideal Cycle Time"]
        * df["Total Count"]
        / df["Operating Time"]
    )

    # --------------------------------------------------------
    # Quality
    # --------------------------------------------------------

    df["Quality"] = (
        df["Good Count"]
        / df["Total Count"]
    )

    # --------------------------------------------------------
    # OEE
    # --------------------------------------------------------

    df["OEE"] = (
        df["Availability"]
        * df["Performance"]
        * df["Quality"]
    )

    # --------------------------------------------------------
    # Throughput
    # --------------------------------------------------------

    df["Throughput"] = df["Good Count"]

    # --------------------------------------------------------
    # Theoretical Capacity
    # --------------------------------------------------------

    df["Theoretical Capacity"] = (
        df["Planned Production Time"]
        / df["Ideal Cycle Time"]
    )

    # --------------------------------------------------------
    # Capacity Utilization
    # --------------------------------------------------------

    df["Capacity Utilization"] = (
        df["Good Count"]
        / df["Theoretical Capacity"]
    )

    return df

# ============================================================
# WORKSTATION ANALYSIS
# ============================================================

def analyze_workstations(production_df):
    """
    Aggregate the 30-day production data by workstation.
    """

    workstation_analysis = (
        production_df
        .groupby("Workstation")
        .agg(
            Total_Production=("Total Count", "sum"),
            Good_Units=("Good Count", "sum"),
            Rejected_Units=("Rejected Count", "sum"),
            Total_Downtime=("Downtime", "sum"),
            Average_Availability=("Availability", "mean"),
            Average_Performance=("Performance", "mean"),
            Average_Quality=("Quality", "mean"),
            Average_OEE=("OEE", "mean"),
            Average_Throughput=("Throughput", "mean"),
            Average_Capacity_Utilization=(
                "Capacity Utilization",
                "mean"
            ),
            Ideal_Cycle_Time=("Ideal Cycle Time", "first"),
            Theoretical_Capacity=("Theoretical Capacity", "first")
        )
        .reset_index()
    )

    # Effective daily good-unit capacity
    workstation_analysis["Effective Daily Capacity"] = (
        workstation_analysis["Good_Units"] / NUMBER_OF_DAYS
    )

    return workstation_analysis

# ============================================================
# BOTTLENECK IDENTIFICATION
# ============================================================

def identify_bottleneck(workstation_analysis):
    """
    Identify the bottleneck using effective daily capacity.

    The workstation with the lowest effective capacity
    is considered the primary bottleneck.
    """

    bottleneck_row = workstation_analysis.loc[
        workstation_analysis["Effective Daily Capacity"].idxmin()
    ]

    return bottleneck_row

# ============================================================
# PRODUCTION LINE THROUGHPUT
# ============================================================

def calculate_line_throughput(production_df):
    """
    Calculate finished-good production line throughput.

    For each day, the line output is constrained by the
    workstation with the lowest good-unit capacity.
    """

    daily_output = (
        production_df
        .groupby("Day")["Good Count"]
        .min()
    )

    total_line_throughput = daily_output.sum()

    average_daily_throughput = daily_output.mean()

    return (
        total_line_throughput,
        average_daily_throughput,
        daily_output
    )

# ============================================================
# OPTIMIZATION SCENARIOS
# ============================================================

def run_optimization_scenarios(bottleneck_name):
    """
    Run baseline and three optimization scenarios.

    Scenario 1:
        10% bottleneck cycle-time reduction.

    Scenario 2:
        20% bottleneck downtime reduction.

    Scenario 3:
        Both improvements combined.
    """

    scenarios = {}

    # --------------------------------------------------------
    # Scenario 0 - Baseline
    # --------------------------------------------------------

    np.random.seed(RANDOM_SEED)

    baseline_df = generate_production_data()

    baseline_df = calculate_kpis(
        baseline_df
    )

    scenarios["Baseline"] = baseline_df

    # --------------------------------------------------------
    # Scenario 1 - Cycle Time Improvement
    # --------------------------------------------------------

    np.random.seed(RANDOM_SEED)

    cycle_time_df = generate_production_data(
        cycle_time_reduction=0.10,
        target_workstation=bottleneck_name
    )

    cycle_time_df = calculate_kpis(
        cycle_time_df
    )

    scenarios["Cycle-Time Improvement"] = cycle_time_df

    # --------------------------------------------------------
    # Scenario 2 - Downtime Reduction
    # --------------------------------------------------------

    np.random.seed(RANDOM_SEED)

    downtime_df = generate_production_data(
        downtime_reduction=0.20,
        target_workstation=bottleneck_name
    )

    downtime_df = calculate_kpis(
        downtime_df
    )

    scenarios["Downtime Reduction"] = downtime_df

    # --------------------------------------------------------
    # Scenario 3 - Combined Improvement
    # --------------------------------------------------------

    np.random.seed(RANDOM_SEED)

    combined_df = generate_production_data(
        cycle_time_reduction=0.10,
        downtime_reduction=0.20,
        target_workstation=bottleneck_name
    )

    combined_df = calculate_kpis(
        combined_df
    )

    scenarios["Combined Improvement"] = combined_df

    return scenarios

def summarize_scenario(
    scenario_df,
    bottleneck_name
):
    """
    Calculate overall metrics for one optimization scenario.
    """

    bottleneck_data = scenario_df[
        scenario_df["Workstation"]
        == bottleneck_name
    ]

    (
        line_throughput,
        average_daily_throughput,
        _
    ) = calculate_line_throughput(
        scenario_df
    )

    return {
        "Line Throughput":
            line_throughput,

        "Average Daily Throughput":
            average_daily_throughput,

        "Total Downtime":
            scenario_df["Downtime"].sum(),

        "Average OEE":
            scenario_df["OEE"].mean(),

        "Bottleneck Good Units":
            bottleneck_data["Good Count"].sum(),

        "Bottleneck Average OEE":
            bottleneck_data["OEE"].mean(),

        "Bottleneck Average Utilization":
            bottleneck_data[
                "Capacity Utilization"
            ].mean()
    }

def create_optimization_summary(
    scenarios,
    bottleneck_name
):
    """
    Create a comparison table for all scenarios.
    """

    results = []

    for scenario_name, scenario_df in scenarios.items():

        metrics = summarize_scenario(
            scenario_df,
            bottleneck_name
        )

        metrics["Scenario"] = scenario_name

        results.append(metrics)

    summary_df = pd.DataFrame(results)

    columns = [
        "Scenario",
        "Line Throughput",
        "Average Daily Throughput",
        "Total Downtime",
        "Average OEE",
        "Bottleneck Good Units",
        "Bottleneck Average OEE",
        "Bottleneck Average Utilization"
    ]

    return summary_df[columns]

def calculate_improvements(optimization_summary):
    """
    Calculate percentage improvement of each scenario
    relative to the baseline.
    """

    baseline = optimization_summary.iloc[0]

    results = optimization_summary.copy()

    results["Throughput Improvement %"] = (
        (
            results["Line Throughput"]
            - baseline["Line Throughput"]
        )
        / baseline["Line Throughput"]
        * 100
    )

    results["OEE Improvement %"] = (
        (
            results["Average OEE"]
            - baseline["Average OEE"]
        )
        / baseline["Average OEE"]
        * 100
    )

    results["Downtime Reduction %"] = (
        (
            baseline["Total Downtime"]
            - results["Total Downtime"]
        )
        / baseline["Total Downtime"]
        * 100
    )

    results["Bottleneck Output Improvement %"] = (
        (
            results["Bottleneck Good Units"]
            - baseline["Bottleneck Good Units"]
        )
        / baseline["Bottleneck Good Units"]
        * 100
    )

    return results

# ============================================================
# VISUALIZATION
# ============================================================

def create_visualizations(
    workstation_analysis,
    optimization_summary,
    improvement_summary
):
    """
    Generate and save project visualizations.
    """

    # --------------------------------------------------------
    # Create results directory
    # --------------------------------------------------------

    import os

    os.makedirs("results", exist_ok=True)

    # --------------------------------------------------------
    # 1. OEE BY WORKSTATION
    # --------------------------------------------------------

    plt.figure(figsize=(9, 5))

    plt.bar(
        workstation_analysis["Workstation"],
        workstation_analysis["Average_OEE"] * 100
    )

    plt.title("Average OEE by Workstation")
    plt.xlabel("Workstation")
    plt.ylabel("OEE (%)")
    plt.ylim(0, 100)

    plt.tight_layout()

    plt.savefig(
        "results/oee_by_workstation.png",
        dpi=300
    )

    plt.close()

    # --------------------------------------------------------
    # 2. THROUGHPUT BY WORKSTATION
    # --------------------------------------------------------

    plt.figure(figsize=(9, 5))

    plt.bar(
        workstation_analysis["Workstation"],
        workstation_analysis["Average_Throughput"]
    )

    plt.title("Average Workstation Throughput")
    plt.xlabel("Workstation")
    plt.ylabel("Good Units per Day")

    plt.tight_layout()

    plt.savefig(
        "results/throughput_analysis.png",
        dpi=300
    )

    plt.close()

    # --------------------------------------------------------
    # 3. BASELINE VS OPTIMIZED THROUGHPUT
    # --------------------------------------------------------

    throughput_scenarios = optimization_summary[
        optimization_summary["Scenario"].isin(
            [
                "Baseline",
                "Cycle-Time Improvement",
                "Downtime Reduction",
                "Combined Improvement"
            ]
        )
    ]

    plt.figure(figsize=(9, 5))

    plt.bar(
        throughput_scenarios["Scenario"],
        throughput_scenarios["Line Throughput"]
    )

    plt.title(
        "Production Line Throughput: "
        "Baseline vs Optimization"
    )

    plt.xlabel("Scenario")
    plt.ylabel("Total Good Units")

    plt.xticks(rotation=15)

    plt.tight_layout()

    plt.savefig(
        "results/before_after_throughput.png",
        dpi=300
    )

    plt.close()

    # --------------------------------------------------------
    # 4. BASELINE VS OPTIMIZED OEE
    # --------------------------------------------------------

    plt.figure(figsize=(9, 5))

    plt.bar(
        improvement_summary["Scenario"],
        improvement_summary["Average OEE"] * 100
    )

    plt.title(
        "Average OEE: Baseline vs Optimization"
    )

    plt.xlabel("Scenario")
    plt.ylabel("Average OEE (%)")

    plt.ylim(0, 100)

    plt.xticks(rotation=15)

    plt.tight_layout()

    plt.savefig(
        "results/before_after_oee.png",
        dpi=300
    )

    plt.close()

    # --------------------------------------------------------
    # 5. DOWNTIME BY WORKSTATION
    # --------------------------------------------------------

    plt.figure(figsize=(9, 5))

    plt.bar(
        workstation_analysis["Workstation"],
        workstation_analysis["Total_Downtime"]
    )

    plt.title("Total Downtime by Workstation")
    plt.xlabel("Workstation")
    plt.ylabel("Downtime (minutes)")

    plt.tight_layout()

    plt.savefig(
        "results/downtime_analysis.png",
        dpi=300
    )

    plt.close()

    print("\nVisualizations generated successfully.")

    # ============================================================
# EXCEL REPORT
# ============================================================

def create_excel_report(
    production_df,
    workstation_analysis,
    optimization_summary,
    improvement_summary,
    bottleneck_name
):
    """
    Create an Excel workbook containing production analysis,
    OEE analysis, optimization results and dashboard data.
    """

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # --------------------------------------------------------
    # Create workbook
    # --------------------------------------------------------

    workbook = Workbook()

    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Create required sheets
    raw_sheet = workbook.create_sheet("Raw Data")
    workstation_sheet = workbook.create_sheet(
        "Workstation Analysis"
    )
    oee_sheet = workbook.create_sheet("OEE Analysis")
    optimization_sheet = workbook.create_sheet(
        "Optimization"
    )
    dashboard_sheet = workbook.create_sheet("Dashboard")

    # --------------------------------------------------------
    # Helper function
    # --------------------------------------------------------

    def write_dataframe(sheet, dataframe):

        # Header
        for column_index, column_name in enumerate(
            dataframe.columns,
            start=1
        ):
            cell = sheet.cell(
                row=1,
                column=column_index
            )

            cell.value = column_name
            cell.font = Font(bold=True)

        # Data
        for row_index, row in enumerate(
            dataframe.itertuples(index=False),
            start=2
        ):

            for column_index, value in enumerate(
                row,
                start=1
            ):

                sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value
                )

        # Adjust column widths
        for column_index in range(
            1,
            len(dataframe.columns) + 1
        ):

            column_letter = get_column_letter(
                column_index
            )

            sheet.column_dimensions[
                column_letter
            ].width = 18

        # Freeze header
        sheet.freeze_panes = "A2"

    # --------------------------------------------------------
    # 1. RAW DATA
    # --------------------------------------------------------

    write_dataframe(
        raw_sheet,
        production_df
    )

    # --------------------------------------------------------
    # 2. WORKSTATION ANALYSIS
    # --------------------------------------------------------

    write_dataframe(
        workstation_sheet,
        workstation_analysis
    )

    # --------------------------------------------------------
    # 3. OEE ANALYSIS
    # --------------------------------------------------------

    oee_columns = [
        "Workstation",
        "Average_OEE",
        "Average_Capacity_Utilization",
        "Average_Throughput",
        "Total_Downtime"
    ]

    oee_data = workstation_analysis[
        oee_columns
    ].copy()

    write_dataframe(
        oee_sheet,
        oee_data
    )

    # --------------------------------------------------------
    # 4. OPTIMIZATION
    # --------------------------------------------------------

    optimization_data = improvement_summary.copy()

    write_dataframe(
        optimization_sheet,
        optimization_data
    )

    # --------------------------------------------------------
    # 5. DASHBOARD
    # --------------------------------------------------------

    dashboard_sheet["A1"] = (
        "MANUFACTURING PRODUCTION LINE OPTIMIZATION"
    )

    dashboard_sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    dashboard_sheet["A3"] = "Key Metric"
    dashboard_sheet["B3"] = "Value"

    dashboard_sheet["A3"].font = Font(
        bold=True
    )

    dashboard_sheet["B3"].font = Font(
        bold=True
    )

    # --------------------------------------------------------
    # Baseline metrics
    # --------------------------------------------------------

    baseline = optimization_summary.iloc[0]

    combined = optimization_summary[
        optimization_summary["Scenario"]
        == "Combined Improvement"
    ].iloc[0]

    baseline_throughput = (
        baseline["Line Throughput"]
    )

    optimized_throughput = (
        combined["Line Throughput"]
    )

    throughput_improvement = (
        (
            optimized_throughput
            - baseline_throughput
        )
        / baseline_throughput
        * 100
    )

    dashboard_data = [
        (
            "Total Production Throughput",
            baseline_throughput
        ),
        (
            "Baseline Average OEE",
            baseline["Average OEE"]
        ),
        (
            "Bottleneck Workstation",
            bottleneck_name
        ),
        (
            "Optimized Throughput",
            optimized_throughput
        ),
        (
            "Throughput Improvement (%)",
            throughput_improvement
        ),
        (
            "Total Baseline Downtime (min)",
            baseline["Total Downtime"]
        ),
        (
            "Optimized Downtime (min)",
            combined["Total Downtime"]
        )
    ]

    for row_index, (metric, value) in enumerate(
        dashboard_data,
        start=4
    ):

        dashboard_sheet.cell(
            row=row_index,
            column=1,
            value=metric
        )

        dashboard_sheet.cell(
            row=row_index,
            column=2,
            value=value
        )

    # --------------------------------------------------------
    # Formatting
    # --------------------------------------------------------

    dashboard_sheet.column_dimensions[
        "A"
    ].width = 35

    dashboard_sheet.column_dimensions[
        "B"
    ].width = 25

    for row in dashboard_sheet.iter_rows(
        min_row=3,
        max_row=10,
        min_col=1,
        max_col=2
    ):

        for cell in row:

            cell.alignment = Alignment(
                horizontal="left"
            )

    # Percentage formatting
    dashboard_sheet["B5"].number_format = "0.00%"
    
    # Save workbook
    workbook.save(
        "production_analysis.xlsx"
    )

    print(
        "\nExcel report created successfully: "
        "production_analysis.xlsx"
    )

# ============================================================
# FORMAT EXCEL DASHBOARD
# ============================================================

# ============================================================
# FORMAT EXCEL DASHBOARD
# ============================================================

def format_excel_dashboard(
    optimization_summary,
    bottleneck_name
):
    """
    Format the Excel dashboard using actual calculated
    production and optimization results.

    The bottleneck and KPI values are passed dynamically
    from the Python analysis.
    """

    from openpyxl import load_workbook
    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side
    )
    from openpyxl.chart import BarChart, Reference

    # --------------------------------------------------------
    # Load existing workbook
    # --------------------------------------------------------

    workbook = load_workbook(
        "production_analysis.xlsx"
    )

    dashboard = workbook["Dashboard"]

    # --------------------------------------------------------
    # Get actual scenario results
    # --------------------------------------------------------

    baseline = optimization_summary[
        optimization_summary["Scenario"] == "Baseline"
    ].iloc[0]

    combined = optimization_summary[
        optimization_summary["Scenario"]
        == "Combined Improvement"
    ].iloc[0]

    # --------------------------------------------------------
    # Calculate actual improvements
    # --------------------------------------------------------

    baseline_throughput = baseline[
        "Line Throughput"
    ]

    optimized_throughput = combined[
        "Line Throughput"
    ]

    throughput_improvement = (
        (
            optimized_throughput
            - baseline_throughput
        )
        / baseline_throughput
    )

    baseline_oee = baseline[
        "Average OEE"
    ]

    optimized_oee = combined[
        "Average OEE"
    ]

    downtime_reduction = (
        (
            baseline["Total Downtime"]
            - combined["Total Downtime"]
        )
        / baseline["Total Downtime"]
    )

    # --------------------------------------------------------
    # Clear existing dashboard
    # --------------------------------------------------------

    dashboard.delete_rows(
        1,
        dashboard.max_row
    )

    # --------------------------------------------------------
    # Colors and styles
    # --------------------------------------------------------

    dark_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    light_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )

    green_fill = PatternFill(
        fill_type="solid",
        fgColor="E2F0D9"
    )

    orange_fill = PatternFill(
        fill_type="solid",
        fgColor="FCE4D6"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    title_font = Font(
        color="FFFFFF",
        bold=True,
        size=20
    )

    section_font = Font(
        color="FFFFFF",
        bold=True,
        size=14
    )

    value_font = Font(
        bold=True,
        size=16
    )

    header_font = Font(
        bold=True
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # --------------------------------------------------------
    # DASHBOARD TITLE
    # --------------------------------------------------------

    dashboard.merge_cells(
        "A1:H2"
    )

    dashboard["A1"] = (
        "MANUFACTURING PRODUCTION LINE "
        "OPTIMIZATION DASHBOARD"
    )

    dashboard["A1"].font = title_font
    dashboard["A1"].fill = dark_fill
    dashboard["A1"].alignment = center_alignment

    dashboard.row_dimensions[1].height = 30
    dashboard.row_dimensions[2].height = 30

    # --------------------------------------------------------
    # KPI CARD 1 - BOTTLENECK
    # --------------------------------------------------------

    dashboard["A4"] = "BOTTLENECK"

    dashboard["A5"] = bottleneck_name

    # --------------------------------------------------------
    # KPI CARD 2 - BASELINE THROUGHPUT
    # --------------------------------------------------------

    dashboard["D4"] = "BASELINE THROUGHPUT"

    dashboard["D5"] = baseline_throughput

    # --------------------------------------------------------
    # KPI CARD 3 - OPTIMIZED THROUGHPUT
    # --------------------------------------------------------

    dashboard["G4"] = "OPTIMIZED THROUGHPUT"

    dashboard["G5"] = optimized_throughput

    # --------------------------------------------------------
    # KPI CARD 4 - THROUGHPUT IMPROVEMENT
    # --------------------------------------------------------

    dashboard["A7"] = "THROUGHPUT IMPROVEMENT"

    dashboard["A8"] = throughput_improvement

    # --------------------------------------------------------
    # KPI CARD 5 - BASELINE OEE
    # --------------------------------------------------------

    dashboard["D7"] = "BASELINE OEE"

    dashboard["D8"] = baseline_oee

    # --------------------------------------------------------
    # KPI CARD 6 - OPTIMIZED OEE
    # --------------------------------------------------------

    dashboard["G7"] = "OPTIMIZED OEE"

    dashboard["G8"] = optimized_oee

    # --------------------------------------------------------
    # KPI CARD 7 - DOWNTIME REDUCTION
    # --------------------------------------------------------

    dashboard["A10"] = "DOWNTIME REDUCTION"

    dashboard["A11"] = downtime_reduction

    # --------------------------------------------------------
    # Style KPI cards
    # --------------------------------------------------------

    card_pairs = [
        ("A4", "A5"),
        ("D4", "D5"),
        ("G4", "G5"),
        ("A7", "A8"),
        ("D7", "D8"),
        ("G7", "G8"),
        ("A10", "A11")
    ]

    for label_cell, value_cell in card_pairs:

        dashboard[label_cell].fill = dark_fill
        dashboard[label_cell].font = white_font
        dashboard[label_cell].alignment = center_alignment
        dashboard[label_cell].border = thin_border

        dashboard[value_cell].fill = light_fill
        dashboard[value_cell].font = value_font
        dashboard[value_cell].alignment = center_alignment
        dashboard[value_cell].border = thin_border

    # --------------------------------------------------------
    # Number formatting
    # --------------------------------------------------------

    dashboard["A8"].number_format = "0.00%"
    dashboard["D8"].number_format = "0.00%"
    dashboard["G8"].number_format = "0.00%"
    dashboard["A11"].number_format = "0.00%"

    dashboard["D5"].number_format = "0"
    dashboard["G5"].number_format = "0"

    # --------------------------------------------------------
    # OPTIMIZATION COMPARISON SECTION
    # --------------------------------------------------------

    dashboard.merge_cells(
        "A14:H14"
    )

    dashboard["A14"] = (
        "OPTIMIZATION SCENARIO COMPARISON"
    )

    dashboard["A14"].font = section_font
    dashboard["A14"].fill = dark_fill
    dashboard["A14"].alignment = center_alignment

    # --------------------------------------------------------
    # Table headers
    # --------------------------------------------------------

    headers = [
        "Scenario",
        "Line Throughput",
        "Average Daily Throughput",
        "Total Downtime",
        "Average OEE",
        "Bottleneck Good Units",
        "Bottleneck OEE",
        "Bottleneck Utilization"
    ]

    for column_index, header in enumerate(
        headers,
        start=1
    ):

        cell = dashboard.cell(
            row=15,
            column=column_index
        )

        cell.value = header
        cell.font = header_font
        cell.fill = light_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # --------------------------------------------------------
    # Write actual scenario data
    # --------------------------------------------------------

    scenarios = [
        "Baseline",
        "Cycle-Time Improvement",
        "Downtime Reduction",
        "Combined Improvement"
    ]

    for row_index, scenario_name in enumerate(
        scenarios,
        start=16
    ):

        scenario_rows = optimization_summary[
            optimization_summary["Scenario"]
            == scenario_name
        ]

        if scenario_rows.empty:
            continue

        scenario = scenario_rows.iloc[0]

        values = [
            scenario["Scenario"],
            scenario["Line Throughput"],
            scenario["Average Daily Throughput"],
            scenario["Total Downtime"],
            scenario["Average OEE"],
            scenario["Bottleneck Good Units"],
            scenario["Bottleneck Average OEE"],
            scenario["Bottleneck Average Utilization"]
        ]

        for column_index, value in enumerate(
            values,
            start=1
        ):

            cell = dashboard.cell(
                row=row_index,
                column=column_index
            )

            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment

    # --------------------------------------------------------
    # Format percentage columns
    # --------------------------------------------------------

    for row in range(16, 20):

        dashboard.cell(
            row=row,
            column=5
        ).number_format = "0.00%"

        dashboard.cell(
            row=row,
            column=7
        ).number_format = "0.00%"

        dashboard.cell(
            row=row,
            column=8
        ).number_format = "0.00%"

    # --------------------------------------------------------
    # Highlight combined improvement row
    # --------------------------------------------------------

    for column in range(1, 9):

        dashboard.cell(
            row=19,
            column=column
        ).fill = green_fill

        dashboard.cell(
            row=19,
            column=column
        ).font = Font(
            bold=True
        )

    # --------------------------------------------------------
    # THROUGHPUT CHART
    # --------------------------------------------------------

    throughput_chart = BarChart()

    throughput_chart.type = "col"

    throughput_chart.title = (
        "Production Throughput by Scenario"
    )

    throughput_chart.y_axis.title = (
        "Units"
    )

    throughput_chart.x_axis.title = (
        "Optimization Scenario"
    )

    throughput_data = Reference(
        dashboard,
        min_col=2,
        min_row=15,
        max_row=19
    )

    scenario_categories = Reference(
        dashboard,
        min_col=1,
        min_row=16,
        max_row=19
    )

    throughput_chart.add_data(
        throughput_data,
        titles_from_data=True
    )

    throughput_chart.set_categories(
        scenario_categories
    )

    throughput_chart.height = 7
    throughput_chart.width = 13

    dashboard.add_chart(
        throughput_chart,
        "A22"
    )

    # --------------------------------------------------------
    # OEE CHART
    # --------------------------------------------------------

    oee_chart = BarChart()

    oee_chart.type = "col"

    oee_chart.title = (
        "OEE by Optimization Scenario"
    )

    oee_chart.y_axis.title = (
        "OEE"
    )

    oee_chart.x_axis.title = (
        "Optimization Scenario"
    )

    oee_data = Reference(
        dashboard,
        min_col=5,
        min_row=15,
        max_row=19
    )

    oee_chart.add_data(
        oee_data,
        titles_from_data=True
    )

    oee_chart.set_categories(
        scenario_categories
    )

    oee_chart.height = 7
    oee_chart.width = 13

    dashboard.add_chart(
        oee_chart,
        "J22"
    )

    # --------------------------------------------------------
    # Column widths
    # --------------------------------------------------------

    column_widths = {
        "A": 28,
        "B": 20,
        "C": 24,
        "D": 22,
        "E": 18,
        "F": 24,
        "G": 20,
        "H": 24
    }

    for column, width in column_widths.items():

        dashboard.column_dimensions[
            column
        ].width = width

    # --------------------------------------------------------
    # Freeze panes
    # --------------------------------------------------------

    dashboard.freeze_panes = "A15"

    # --------------------------------------------------------
    # Save workbook
    # --------------------------------------------------------

    workbook.save(
        "production_analysis.xlsx"
    )

    print(
        "\nExcel dashboard formatted successfully."
    )


# MAIN PROGRAM
# ============================================================
# DISPLAY PROJECT CONFIGURATION
# ============================================================

def display_configuration():

    print("=" * 70)
    print("MANUFACTURING PRODUCTION LINE OPTIMIZATION")
    print("=" * 70)

    print("\nProduction Line:")
    print(
        "Raw Material → Cutting → Machining → "
        "Assembly → Inspection → Packing"
    )

    print("\nSimulation Period:")
    print("30 production days")

    print("\nShift Duration:")
    print("480 minutes (8-hour shift)")

    print("\nKey Manufacturing KPIs:")
    print(
        "Availability, Performance, Quality, OEE, "
        "Throughput, Capacity Utilization"
    )

    print("\nOptimization Scenarios:")
    print("1. Baseline")
    print("2. Bottleneck Cycle-Time Improvement")
    print("3. Bottleneck Downtime Reduction")
    print("4. Combined Improvement")

    print("\nBottleneck Identification:")
    print(
        "Automatically determined from workstation "
        "effective capacity."
    )

    print("=" * 70)

if __name__ == "__main__":

    # ========================================================
    # PROJECT CONFIGURATION
    # ========================================================

    display_configuration()

    # ========================================================
    # 1. GENERATE RAW PRODUCTION DATA
    # ========================================================

    production_df = generate_production_data()

    print("\nProduction data generated successfully.")

    print("\nRaw dataset shape:")
    print(production_df.shape)

    # ========================================================
    # 2. CALCULATE MANUFACTURING KPIs
    # ========================================================

    production_df = calculate_kpis(
        production_df
    )

    print("\nKPI calculations completed.")

    # Save processed production data
    production_df.to_csv(
        "production_data.csv",
        index=False
    )

    print(
        "Production dataset saved to "
        "production_data.csv"
    )

    # ========================================================
    # 3. WORKSTATION ANALYSIS
    # ========================================================

    workstation_analysis = analyze_workstations(
        production_df
    )

    print("\n" + "=" * 70)
    print("30-DAY WORKSTATION ANALYSIS")
    print("=" * 70)

    print(
        workstation_analysis[
            [
                "Workstation",
                "Good_Units",
                "Total_Downtime",
                "Average_OEE",
                "Average_Throughput",
                "Average_Capacity_Utilization",
                "Effective Daily Capacity"
            ]
        ].to_string(index=False)
    )

    # ========================================================
    # 4. BOTTLENECK IDENTIFICATION
    # ========================================================

    bottleneck = identify_bottleneck(
        workstation_analysis
    )

    print("\n" + "=" * 70)
    print("BOTTLENECK IDENTIFICATION")
    print("=" * 70)

    print(
        f"Identified bottleneck: "
        f"{bottleneck['Workstation']}"
    )

    print(
        f"Effective daily capacity: "
        f"{bottleneck['Effective Daily Capacity']:.2f} "
        f"units/day"
    )

    print(
        f"Average OEE: "
        f"{bottleneck['Average_OEE']:.2%}"
    )

    print(
        f"Average capacity utilization: "
        f"{bottleneck['Average_Capacity_Utilization']:.2%}"
    )

    print(
        f"Total downtime: "
        f"{bottleneck['Total_Downtime']:.2f} minutes"
    )

    # ========================================================
    # 5. OPTIMIZATION SCENARIOS
    # ========================================================

    print("\n" + "=" * 70)
    print("OPTIMIZATION SCENARIOS")
    print("=" * 70)

    scenarios = run_optimization_scenarios(
        bottleneck["Workstation"]
    )

    optimization_summary = create_optimization_summary(
        scenarios,
        bottleneck["Workstation"]
    )

    print(
        optimization_summary.to_string(
            index=False
        )
    )

    # ========================================================
    # 6. CALCULATE IMPROVEMENTS
    # ========================================================

    improvement_summary = calculate_improvements(
        optimization_summary
    )

    print("\n" + "=" * 70)
    print("OPTIMIZATION IMPROVEMENTS")
    print("=" * 70)

    print(
        improvement_summary[
            [
                "Scenario",
                "Throughput Improvement %",
                "OEE Improvement %",
                "Downtime Reduction %",
                "Bottleneck Output Improvement %"
            ]
        ].to_string(index=False)
    )

    # ========================================================
    # 7. CREATE VISUALIZATIONS
    # ========================================================

    create_visualizations(
        workstation_analysis,
        optimization_summary,
        improvement_summary
    )

        # ========================================================
    # 8. CREATE EXCEL REPORT
    # ========================================================

    create_excel_report(
        production_df,
        workstation_analysis,
        optimization_summary,
        improvement_summary,
        bottleneck["Workstation"]
    )

    format_excel_dashboard(
    optimization_summary,
    bottleneck["Workstation"]
)
    
    print("\n" + "=" * 70)
    print("PROJECT ANALYSIS COMPLETED")
    print("=" * 70)